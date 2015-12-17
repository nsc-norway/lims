import sys
import itertools
from openpyxl import load_workbook
import StringIO
import requests
import datetime
from genologics.lims import *
from genologics import config

# SwissLab data import script

# Use:
# python swisslab-import.py PROCESS-ID SWL-FILE
# Arguments:
#  PROCESS-ID: LIMS-ID of a process on which this script is run.
#  SWL-FILE:   LIMS-ID of Artifact object for Excel file

REQUIRED_COLUMNS = {
        "Test SWL Diag": str,
        "Referral reason Diag": str,
        "Archive position Diag": str,
        "Alternative sample ID Diag": str,
        "Gene panel Diag": str,
        "Kit version Diag": str,
        "Analysis type Diag": str
        }

OPTIONAL_COLUMNS = {
        "Analysis registered Diag": datetime.datetime.date
        }

TRIO_COLUMNS = {
        "Family number Diag": str,
        "Relation Diag": str,
        "Sex Diag": str
        }

ALL_COLUMNS = frozenset(REQUIRED_COLUMNS.keys() + OPTIONAL_COLUMNS.keys() + TRIO_COLUMNS.keys())

def get_swl_data(filename):
    """Parsing of sample information table in Excel format.
    
    Returns a dict indexed by sample name, where each value is a list 
    of (udfname, udfvalue) tuples.
    { sample name => [ (udfname, udfvalue), ... ] }
    """
    try:
        wb = load_workbook(filename, data_only=True)
    except IOError:
        print "Cannot read the SwissLab file, make sure it is in Excel format"
        sys.exit(1)

    ws = wb['Samples']

    assert ws['A1'].value == "Sample/Name"
    # Get { header name => column index }
    headers = {}
    for i in itertools.count(2):
        h = ws.cell(column=i, row=1).value
        if h:
            if h in ALL_COLUMNS:
                headers[h] = i
        else:
            break

    missing = ALL_COLUMNS - frozenset(headers.keys())
    if missing:
        print "Missing the following columns in the Excel file:", ",".join(missing)
        sys.exit(1)

    data = {}
    for row in itertools.count(2):
        sample_name = ws.cell(column=1, row=row).value
        if sample_name:
            sample_name = str(sample_name)

            sample_data = get_all_fields(
                    sample_name,
                    ws, row, headers,
                    REQUIRED_COLUMNS, 
                    True
                    )

            sample_data += get_all_fields(
                    sample_name,
                    ws, row, headers,
                    OPTIONAL_COLUMNS, 
                    False
                    )

            trio_fields_required = any(
                    val == "trio" and name == "Analysis type Diag"
                    for name, val in sample_data
                    )
            sample_data += get_all_fields(
                    sample_name,
                    ws, row, headers,
                    TRIO_COLUMNS,
                    trio_fields_required
                    )
            data[sample_name] = sample_data
        else:
            break

    return data


def get_all_fields(sample_name, ws, row, headers, columns, required):
    """Get a list of tuples, one for each column in the columns parameter."""
    result = []
    for col, value_type in columns.items():
        v = ws.cell(column=headers[col], row=row).value
        if not v is None:
            try:
                result.append((col, value_type(v)))
            except ValueError:
                print "Cannot convert '" +  str(v) + "' to a", value_type,\
                        "for sample", sample_name, ", column", col
                sys.exit(1)
        elif required:
            print "Missing required value for '", col, "' for sample", sample_name
            sys.exit(1)
    return result
            

def main(process_id, swl_file_id, ignore_duplicates=False):
    lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)

    swl_file_art_obj = Artifact(lims, id=swl_file_id)
    file_data = None
    try:
        if len(swl_file_art_obj.files) == 1:
            file_data = swl_file_art_obj.files[0].download()
    except requests.exceptions.HTTPError:
        pass

    if file_data is None:
        print "Could not access the SwissLab file, check that it has been uploaded"
        return # This is not a fatal error: some projects may not have SwissLab files

    swl_io_obj = StringIO.StringIO(file_data)
    swisslab_data = get_swl_data(swl_io_obj)

    process = Process(lims, id=process_id)
    inputs = process.all_inputs(unique=True, resolve=True)
    assert all(len(input.samples) == 1 for input in inputs)
    samples = lims.get_batch(input.samples[0] for input in inputs)

    uniques = set(sample.name for sample in samples)
    if len(uniques) < len(samples):
        dupes = []
        for sample in samples:
            try:
                uniques.remove(sample.name)
            except KeyError:
                dupes.append(sample)
        print "Duplicate sample IDs detected for:", ", ".join(dupes)
        sys.exit(1)
    
    for sample in samples:
        try:
            fields = swisslab_data[sample.name]
        except KeyError:
            print "Failed trying to look up sample", sample.name, "in the provided table"
            sys.exit(1)

        for udfname, udfval in fields:
            sample.udf[udfname] = udfval

    # Check for duplicates if requested
    if not ignore_duplicates:
        existing_samples = lims.get_samples(name=[sample.name for sample in samples])
        if len(existing_samples) > len(samples):
            # existings_samples will always be >= samples as a set, because samples are 
            # contained in existing samples.
            # Only inspect further in the rare case when there is a dupe. There is still not 
            # necessarily a duplicate, since even if there's a sample with the same name, it
            # may have a different gene panel.
            lims.get_batch(existing_samples) # This will not refresh the ones which are already
                                                                # cached (and modified)
            try:
                existing_sample_keys = [(sample.name, sample.udf['Gene panel Diag']) for sample in existing_samples]
            except KeyError:
                print "A sample", sample.name, "is already in the system, but has no gene panel"
                sys.exit(1)
            step_sample_keys = ((sample.name, sample.udf['Gene panel Diag']) for sample in samples)
            dupes = []
            # Identify all items that occur more than once in existing samples
            for sample_key in step_sample_keys:
                try:
                    existing_sample_keys.remove(sample_key)
                    existing_sample_keys.remove(sample_key)
                    dupes.append(sample_key)
                except KeyError:
                    pass

            if dupes:
                print "Existing sample(s) with same name and gene panel found for:",
                print ",".join(dupes)
                sys.exit(1)

    lims.put_batch(samples)
    print "Imported data for", len(samples), "samples"


main(*sys.argv[1:])

