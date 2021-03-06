# -*- coding: utf-8 -*-

# Note: Python 3 source code, run under SCL.

# This script is designed to output an Excel file to help dilution for cluster gen.
# Based on: ../sample-prep/dna_norm_worksheet.py

# Generate table for normalisation, for clustering

# Installation of requirements on RHEL6/Centos6:
#$ sudo yum install rh-python35
#$ sudo scl enable rh-python35 bash # This opens a sub-shell, run the following in the sub-shell:
## pip install openpyxl
## pip install requests
# (exit; done)


# The script can be used on different steps, and thus has to support different sets of UDF names.
# The third command line argument is used as a code to select UDFs from which to take the values.

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import Workbook
from genologics.lims import *
from genologics import config
import sys
import re


if len(sys.argv) > 3:
    code = sys.argv[3]
else:
    code = "TruSeq"
if code == "Diag":
    CONC_INPUT_UDF = "Conc Input (nM) Diag"
    VOLUME_PHIX_UDF = "Volume PhiX (uL) Diag"
    VOLUME_FINAL_UDF = "Volume final (uL) Diag"
elif code == "TruSeq":
    CONC_INPUT_UDF = "Conc. Input (nM) TruSeq DNA"
    VOLUME_PHIX_UDF = "Volume PhiX (uL) TruSeq DNA"
    VOLUME_FINAL_UDF = "Volume final (uL) TruSeq DNA"


START_COL = 2
START_ROW = 2

warning = []

lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)
process = Process(lims, id=sys.argv[1])

wb = Workbook()
ws = wb.active
side_style = Side(border_style="thin")
thick_side_style = Side(border_style="thick")

def sort_key(elem):
    input, output = elem
    container, well = output.location
    row, col = well.split(":")
    return (container.id, int(col), row)

formatting = [
    ("Sample name",         16, None),   
    ("Well",                8, None),
    ("Sample conc. [nM]",   10, '0.00'),
    ("Conc. Input [nM]",    10,'0.0'),
    ("Volum final [uL]",    10, '0'),
    ("Input [uL]",          10, '0.00'),
    ("RSB [uL]",            10, '0.00'),
    ("PhiX [uL]",           10, '0.00'),
    ("Well strip",          9,  '0')
    ]

col_index = {name: i for i, (name, _, _) in enumerate(formatting)}


for i, (header, width, number_format) in enumerate(formatting):
    cell = ws.cell(row=START_ROW, column=i+START_COL)
    left_style = thick_side_style if i == 0 else side_style
    right_style = thick_side_style if i == len(formatting)-1 else side_style
    border_style = Border(top=thick_side_style, left=left_style, right=right_style, bottom=side_style)
    cell.border = border_style
    cell.alignment = Alignment(wrapText=True, horizontal='center')
    cell.font = Font(bold=True)
    cell.value = header
    ws.column_dimensions[chr(ord('A')+i+START_COL-1)].width = width

i_os = process.input_output_maps
inputs_outputs = [(i['uri'], o['uri']) for i,o in i_os if o['output-generation-type'] == 'PerInput']
lims.get_batch([i for i,o in inputs_outputs] + [o for i,o in inputs_outputs])

for row_index, (input, output) in enumerate(sorted(inputs_outputs, key=sort_key), 1+START_ROW):

    top_style = thick_side_style if row_index == 1+START_ROW else side_style
    bottom_style = thick_side_style if row_index == START_ROW+len(inputs_outputs) else side_style

    ws.cell(row=row_index, column=START_COL).border =\
            Border(top=top_style, left=thick_side_style, right=side_style, bottom=bottom_style)
    ws.cell(row=row_index, column=START_COL+len(formatting)-1).border =\
            Border(top=top_style, left=side_style, right=thick_side_style, bottom=bottom_style)
    for i in range(START_COL+1, START_COL+len(formatting)-1):
        ws.cell(row=row_index, column=i).border = \
            Border(top=top_style, left=side_style, right=side_style, bottom=bottom_style)

    for i in range(len(formatting)):
        ws.cell(row=row_index,column=i+START_COL).alignment = Alignment(horizontal='center')
        if formatting[i][2]:
            ws.cell(row=row_index,column=i+START_COL).number_format = formatting[i][2]


    # ---- Artifact info ----
    ws.cell(row=row_index, column=START_COL+col_index["Sample name"]).value = input.name
    ws.cell(row=row_index, column=START_COL+col_index["Well"]).value = input.location[1].replace(":", "")
    molarity = input.udf['Molarity']
    ws.cell(row=row_index, column=START_COL+col_index["Sample conc. [nM]"]).value = molarity

    # ---- Parameters (use specific for sample, or default) ----
    try:
        conc_input = output.udf[CONC_INPUT_UDF]
        phix_input = output.udf[VOLUME_PHIX_UDF]
        final_volume = output.udf[VOLUME_FINAL_UDF]
    except KeyError as e:
        print ("Error: missing value for", e, "for sample", output.name)
        sys.exit(1)

    ws.cell(row=row_index, column=START_COL+col_index["Conc. Input [nM]"]).value = conc_input
    ws.cell(row=row_index, column=START_COL+col_index["Volum final [uL]"]).value = final_volume
    ws.cell(row=row_index, column=START_COL+col_index["PhiX [uL]"]).value = phix_input

    # ---- Calculated quantities ----
    
    # Input mirolitres
    conc_input_coord = ws.cell(row=row_index, column=START_COL+col_index["Conc. Input [nM]"]).coordinate
    final_volume_coord = ws.cell(row=row_index, column=START_COL+col_index["Volum final [uL]"]).coordinate
    molarity_coord = ws.cell(row=row_index, column=START_COL+col_index["Sample conc. [nM]"]).coordinate
    ws.cell(row=row_index, column=START_COL+col_index["Input [uL]"]).value = '=(({0}*{1})/{2})'.format(
        conc_input_coord, final_volume_coord, molarity_coord)

    # RSB microlitres
    phix_volume_coord = ws.cell(row=row_index, column=START_COL+col_index["PhiX [uL]"]).coordinate
    input_ul_coord = ws.cell(row=row_index, column=START_COL+col_index["Input [uL]"]).coordinate
    ws.cell(row=row_index, column=START_COL+col_index["RSB [uL]"]).value = "={0}-{1}-{2}".format(
            final_volume_coord, input_ul_coord, phix_volume_coord)

    # ---- Well strip ----
    ws.cell(row=row_index, column=START_COL+col_index["Well strip"]).value = int(output.location[1].partition(':')[0])


lims.put_batch(o for i,o in inputs_outputs)

wb.save(sys.argv[2] + '.xlsx')

if warning:
    print("Warning: too low input concentration for samples:", ", ".join(warning), ".")
    sys.exit(1)

