# -*- coding: utf-8 -*-

# Note: Python 3 source code, run under SCL.

# This script is designed to output an Excel file to help with RNA sample
# prep.

# See the dna equivalent: dna_norm_worksheet.py, for how to install requirements

from openpyxl.styles import Border, Side, Alignment
from openpyxl import Workbook
import openpyxl
from genologics.lims import *
from genologics import config
import sys
import re


warning = []

lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)
process = Process(lims, id=sys.argv[1])

wb = Workbook()
ws = wb.active

side_style = Side(border_style="thin")
border_style = Border(top=side_style, left=side_style, right=side_style, bottom=side_style)

def sort_key(elem):
    input, output = elem
    container, well = output.location
    row, col = well.split(":")
    return (container, int(col), row)

def get_or_default(udfname):
    try:
        return output.udf[udfname]
    except KeyError:
        val = process.udf[udfname]
        output.udf[udfname] = val
        return val

headers = [
    "Project",
    "Sample name",
    "Position",
    "Input (ng)",
    "Vol total (µL)",
    "Volume RNA (µL)",
    "Volume dH2O (µL)",
    "Index name",
    ]

for i, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=i)
    cell.value = header
    ws.column_dimensions[chr(ord('A')+i)].bestFit=True
    ws.column_dimensions[chr(ord('A')+i)].hidden=False
    for i in range(1, len(headers)+1):
        ws.cell(row=1, column=i).border = border_style

for i, width in enumerate([20, 16, 8, 8, 6, 12, 12, 16]):
    ws.column_dimensions[chr(ord('A')+i)].width = width

i_os = process.input_output_maps
inputs_outputs = [(i['uri'], o['uri']) for i,o in i_os if o['output-generation-type'] == 'PerInput']
lims.get_batch([i for i,o in inputs_outputs] + [o for i,o in inputs_outputs])
lims.get_batch(i.samples[0] for i, o in inputs_outputs)
for row_index, (input, output) in enumerate(sorted(inputs_outputs, key=sort_key), 2):
    for i in range(1, len(headers)+1):
        ws.cell(row=row_index, column=i).border = border_style

    # ---- Sample info ----
    ws.cell(row=row_index, column=1).value = input.samples[0].project.name
    ws.cell(row=row_index, column=2).value = input.name
    ws.cell(row=row_index, column=3).value = output.location[1].replace(":", "")

    # ---- Parameters (use specific for sample, or default) ----
    input_ng = get_or_default('Input (ng)')
    total_volume = get_or_default('Volume (uL)')

    ws.cell(row=row_index, column=4).value = input_ng
    ws.cell(row=row_index, column=5).value = total_volume


    # ---- Calculated quantities ----
    try:
        conc = input.samples[0].udf['Sample conc. (ng/ul)']
    except KeyError:
        ws.cell(row=row_index, column=6).value = "MISSING_CONC"
        continue

    if conc == 0.0:
        sample_volume = total_volume + 1
    else:
        sample_volume = input_ng * 1.0 / conc
    buffer_volume = total_volume - sample_volume

    if buffer_volume < 0:
        buffer_volume = 0.0
        sample_volume = total_volume
        warning.append(output.name)

    ws.cell(row=row_index, column=6).value = sample_volume
    ws.cell(row=row_index, column=6).number_format = "0.0"
    ws.cell(row=row_index, column=7).value = buffer_volume
    ws.cell(row=row_index, column=7).number_format = "0.0"

    # ---- Index ----
    ws.cell(row=row_index, column=8).value = re.sub(r" \([ACGT-]+\)$", "", next(iter(output.reagent_labels)))


lims.put_batch(o for i,o in inputs_outputs)

wb.save(sys.argv[2] + '.xlsx')

if warning:
    print("Warning: too low input concentration for samples:", ", ".join(warning), ".")
    sys.exit(1)

