# Note: Python 3 source code, run under SCL.

# -*- coding: utf-8 -*-

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import Workbook
import openpyxl
from genologics.lims import *
from genologics import config
import sys

# Script usage: generate_worksheet.py <PROCESS_ID> <OUTPUT_FILE_ID>

START_SKIP_ROWS = 2

lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)
process = Process(lims, id=sys.argv[1])

wb = Workbook()
ws = wb.active

side_style = Side(border_style="thin")
border_style = Border(top=side_style, left=side_style, right=side_style, bottom=side_style)

def sort_key(elem):
    input, output = elem
    container, well = output.location
    row, col = well.split(":")
    return (container.id, int(col), row)

headers = [
    "Antall",
    "Pos Rack",
    "Rack",
    "Prøvenummer",
    "Arkivposisjon",
    "Alternative Sample ID",
    "Konsentrasjon ng/µL",
    "Posisjon",
    ]

title_cell = ws.cell(row=1, column=1)
title_cell.value = "Oppsett Nextera Flex"
title_cell.font = Font(size=14, bold=True)

for i, header in enumerate(headers, 1):
    cell = ws.cell(row=START_SKIP_ROWS+1, column=i)
    cell.value = header
    ws.column_dimensions[chr(ord('A')+i)].bestFit=True
    ws.column_dimensions[chr(ord('A')+i)].hidden=False
    for i in range(1, len(headers)+1):
        ws.cell(row=START_SKIP_ROWS+1, column=i).border = border_style
        ws.cell(row=START_SKIP_ROWS+1, column=i).font = Font(bold=True)

for i, width in enumerate([7, 7, 7, 32, 12, 18, 8, 6, 8]):
    ws.column_dimensions[chr(ord('A')+i)].width = width

i_os = process.input_output_maps
inputs_outputs = [(i['uri'], o['uri']) for i,o in i_os if o['output-generation-type'] == 'PerInput']
lims.get_batch([i for i,o in inputs_outputs] + [o for i,o in inputs_outputs])
lims.get_batch(i.samples[0] for i, o in inputs_outputs)
for row_index, (input, output) in enumerate(sorted(inputs_outputs, key=sort_key), 2+START_SKIP_ROWS):
    for i in range(1, len(headers)+1):
        ws.cell(row=row_index, column=i).border = border_style
    col = iter(range(1, len(headers)+1))
    well_row, well_col = output.location[1].split(":")
    well_index = (int(well_col) - 1) * 8 + "ABCDEFGH".index(well_row)
    ws.cell(row=row_index, column=next(col)).value = row_index-1-START_SKIP_ROWS
    ws.cell(row=row_index, column=next(col)).value = 1 + well_index % 32
    ws.cell(row=row_index, column=next(col)).value = "Rack{0}".format(1 + well_index // 32)
    ws.cell(row=row_index, column=next(col)).value = input.name
    ws.cell(row=row_index, column=next(col)).value = input.samples[0].udf.get('Archive position Diag', 'UKJENT')
    ws.cell(row=row_index, column=next(col)).value = input.samples[0].udf.get('Alternative Sample ID', 'UKJENT')
    try:
        conc = input.samples[0].udf['Sample conc. (ng/ul)']
    except KeyError:
        ws.cell(row=row_index, column=next(col)).value = "UKJENT"
        continue
    ws.cell(row=row_index, column=next(col)).value = conc
    ws.cell(row=row_index, column=next(col)).value = output.location[1].replace(":", "")
    for i in range(2, len(headers)+1):
        ws.cell(row=row_index, column=i).alignment = Alignment(horizontal="right")

wb.save(sys.argv[2] + '.xlsx')

