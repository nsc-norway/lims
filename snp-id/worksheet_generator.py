# Note: Python 3 source code, run under SCL.

# -*- coding: utf-8 -*-

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import Workbook
import openpyxl
from genologics.lims import *
from genologics import config
import sys

# Script usage: generate_worksheet.py <PROCESS_ID> <OUTPUT_FILE_ID>

START_SKIP_ROWS = 2

lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)
process = Process(lims, id=sys.argv[1])

wb = Workbook()
ws = wb.active

side_style = Side(border_style="thin")
border_style = Border(top=side_style, left=side_style, right=side_style, bottom=side_style)

def sort_key(elem):
    input, output = elem
    container, well = output.location
    row, col = well.split(":")
    return (container, int(col), row)

headers = [
    "Antall",
    "Pos Rack",
    "Rack",
    "Prøvenummer",
    "Arkivposisjon",
    "Alternative Sample ID",
    "Konsentrasjon ng/µL",
    "Posisjon",
    "µL EB",
    "µL DNA",
    ]

title_cell = ws.cell(row=1, column=1)
title_cell.value = "Oppsett til SNP-ID"
title_cell.font = Font(size=14, bold=True)

for i, header in enumerate(headers, 1):
    cell = ws.cell(row=START_SKIP_ROWS+1, column=i)
    cell.value = header
    ws.column_dimensions[chr(ord('A')+i)].bestFit=True
    ws.column_dimensions[chr(ord('A')+i)].hidden=False
    for i in range(1, len(headers)+1):
        ws.cell(row=START_SKIP_ROWS+1, column=i).border = border_style
        ws.cell(row=START_SKIP_ROWS+1, column=i).font = Font(bold=True)

for i, width in enumerate([7, 7, 7, 32, 12, 18, 8, 6, 8]):
    ws.column_dimensions[chr(ord('A')+i)].width = width

i_os = process.input_output_maps
inputs_outputs = [(i['uri'], o['uri']) for i,o in i_os if o['output-generation-type'] == 'PerInput']
lims.get_batch([i for i,o in inputs_outputs] + [o for i,o in inputs_outputs])
lims.get_batch(i.samples[0] for i, o in inputs_outputs)
tube_counter = 0
for row_index, (input, output) in enumerate(sorted(inputs_outputs, key=sort_key), 2+START_SKIP_ROWS):
    for i in range(1, len(headers)+1):
        ws.cell(row=row_index, column=i).border = border_style
    col = iter(range(1, len(headers)+1))
    # Different handling of tube and 96 plate (2D Barcodes tube rack)
    if input.location[0].type_name == '96 well plate':
        labware = "Rack2DTubes"
        position_id = input.location[1].replace(":", "")
    else:
        labware = "Rack%d" % ((tube_counter // 32) + 1)
        position_id = str((tube_counter % 32) + 1)
        tube_counter += 1
    ws.cell(row=row_index, column=next(col)).value = row_index-1-START_SKIP_ROWS
    ws.cell(row=row_index, column=next(col)).value = position_id 
    ws.cell(row=row_index, column=next(col)).value = labware 
    ws.cell(row=row_index, column=next(col)).value = input.name
    ws.cell(row=row_index, column=next(col)).value = input.samples[0].udf.get('Archive position Diag', 'UKJENT')
    ws.cell(row=row_index, column=next(col)).value = input.samples[0].udf.get('Alternative sample ID Diag', 'UKJENT')
    try:
        conc = input.samples[0].udf['Sample conc. (ng/ul)']
    except KeyError:
        ws.cell(row=row_index, column=next(col)).value = "UKJENT"
        continue
    ws.cell(row=row_index, column=next(col)).value = conc
    ws.cell(row=row_index, column=next(col)).value = output.location[1].replace(":", "")
    for i in range(2, len(headers)+1):
        ws.cell(row=row_index, column=i).alignment = Alignment(horizontal="right")
    if conc >= 9 and conc <= 180:
        ws.cell(row=row_index, column=next(col)).value = 43
        ws.cell(row=row_index, column=next(col)).value = 2
    else:
        # Compute 3 ng/uL in 45 uL total volume
        if conc == 0.0:
            sample_vol = 2
        else:
            sample_vol = (3 * 45) / conc
        ws.cell(row=row_index, column=next(col)).value = max(0, 45 - sample_vol)
        ws.cell(row=row_index, column=next(col)).value = sample_vol

wb.save(sys.argv[2] + '.xlsx')

