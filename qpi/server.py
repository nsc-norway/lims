import datetime
import time
import re
import os
import io
import threading
import yaml
import json
import sys
import string
import base64
import csv
import Queue as Mod_Queue # Due to name conflict with genologics
from flask import Flask, url_for, abort, jsonify, Response, request,\
        render_template, redirect
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import requests

from genologics.lims import *
from genologics import config

# External project creation backend server

PROJECT_TYPES = ['FHI-Swift-FileV2', 'MIK-Swift',  'FHI-Swift']

# This will be registered as a WSGI application under a path.

app = Flask(__name__)
lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)

# Global dict that maps type name to job object for project type
# (only one project may be created at a time)
workers = {}
workers_lock = threading.Lock()

def render_index_page(**kwargs):

    project_name_presets = {}
    project_types_for_filename = []
    for pt in PROJECT_TYPES:
        project_def = get_project_def(pt)
        project_name_presets[pt] = project_def.get('project_name_placeholder', '').format(datetime.date.today())
        if project_def.get('add_file_basename_to_project_name'):
            project_types_for_filename.append(pt)

    return render_template("index.html", 
                project_types=PROJECT_TYPES, project_name_presets=project_name_presets,
                project_types_for_filename=project_types_for_filename, **kwargs)

@app.route("/")
def get_project_start_page():
    return render_index_page()


def get_project_def(project_type):
    """Get the project type JSON definition. Calls abort()
    on error, to return HTTP status code.
    
    Returns a dict from deserialized JSON."""
    try:
        project_type_safe = secure_filename(project_type)
        if project_type_safe == "":
            raise ValueError("Invalid project name")
        with open(os.path.join(
                os.path.dirname(__file__),
                "config", project_type_safe + ".yaml")
                ) as f:
            project_data = yaml.safe_load(f)
            return project_data
    except IOError as e:
        abort(500, "Error while getting project type data: " + str(e))


@app.route("/submit", methods=["POST"])
def submit_project():
    """Trigger creation of a project. Instantly returns a
    page showing the status of project creation."""

    template = request.form.get('template', '')
    username = request.form.get('username', '')
    password = request.form.get('password', '')
    projectname = request.form.get('projectname', '')


    if '' in [username, password, template, projectname]:
        return render_index_page(username=username, password=password,
                preset_project_type=template, projectname=projectname,
                error_message= "Please specify username, password, project name and template.")

    if not projectname.replace("-", "").replace("_", "").isalnum():
        return render_index_page(username=username, password=password,
                preset_project_type=template, projectname=projectname,
                error_message= "Project name should only contain alphanumerics and hyphen (-).")

    project_template_data = get_project_def(template)
    
    if projectname == project_template_data.get('project_name_placeholder') or \
            (project_template_data.get('add_file_basename_to_project_name') and \
                     projectname.startswith(project_template_data.get('project_name_placeholder'))):
        return render_index_page(username=username, password=password,
                preset_project_type=template, projectname=projectname,
                error_message= "Please change the project name from the placeholder.")

    try:
        f = request.files['sample_file']
        file_object = io.BytesIO()
        f.save(file_object)
        file_name = secure_filename(f.filename)
        if file_name == "":
            raise ValueError("No file provided")
    except (KeyError, ValueError):
        return render_index_page(username=username, password=password, 
                preset_project_type=template, projectname=projectname,
                error_message="Sample file upload failed. Make sure sample file is specified.")

    with workers_lock:
        worker = workers.setdefault(
                    projectname, ProjectWorker(project_template_data)
                    )
        if not worker.active:
            try:
                worker.start_job(username, password, projectname,
                        file_name, file_object)
            except LimsCredentialsError:
                # Why abort() here, not render_template?: We can't put the
                # file back into the response, so better encourage the user
                # to press back and try again (with the file).
                abort(403, "Incorrect username or password, please go back "
                        "and try again.".format(username, password))
            except Exception as e:
                abort(500, "LIMS seems to be unreachable, or bug in job creation: {0}".format(e))
        
        # Cleanup old workers
        for w in list(workers):
            if workers[w].rundate and workers[w].rundate < datetime.date.today():
                del workers[w]

    return redirect(url_for('get_project_status', projectname=projectname))


@app.route("/status/<projectname>")
def get_project_status(projectname):
    """Status page"""

    parameters = {'evtSourceUrl': url_for('get_stream', projectname=projectname)}
    return render_template("status.html", **parameters)


@app.route("/stream/<projectname>")
def get_stream(projectname):
    """SSE stream with progress."""
    with workers_lock:
        try:
            worker = workers[projectname]
        except KeyError:
            abort(404, "Project is not being processed")
        if worker.job is None:
            abort(500, "No import has been started")
    if worker.job:
        stream = status_stream(worker.job)
        return Response(stream, mimetype="text/event-stream")
    else:
        abort(404, "No job found for this project")


def status_repr(job):
    task_statuses = [
            {
                "running": task.running,
                "error": task.error,
                "completed": task.completed,
                "status": task.status,
                "name": task.NAME
            } for task in job.tasks
        ]
    return json.dumps({
            "project_title": job.projectname,
            "step_url": job.step_url,
            "task_statuses": task_statuses,
            "running": job.running,
            "error": job.error,
            "completed": job.completed
        })


def status_stream(job):
    if job.completed:
        brief_status = json.dumps({
            "project_title": job.projectname,
            "step_url": job.step_url,
            "task_statuses": [],
            "running": False,
            "error": job.error,
            "completed": job.completed
        })
        yield "event: status\ndata: " + brief_status + "\n\n"
        yield "event: shutdown\ndata: null\n\n"
        return
    else:
        yield "event: status\ndata: " + status_repr(job) + "\n\n"

    while job.queue.get(block=True) == "status":
        try:
            # Discard entries if they pile up here
            # Only latest (current) status is of interest
            while job.queue.get(timeout=0) == "status":
                pass
        except Mod_Queue.Empty:
            yield "event: status\ndata: " + status_repr(job) + "\n\n"
        else: # In case we received anything other than "status"
            yield "event: status\ndata: " + status_repr(job) + "\n\n"
            yield "event: shutdown\ndata: null\n\n"
            break


class LimsCredentialsError(ValueError):
    pass


class Task(object):
    NAME = None

    def __init__(self, job):
        self.running = False
        self.completed = False
        self._status = None
        self.error = False
        self.job = job

    def __call__(self):
        try:
            self.running = True
            self.status = None
            self.run()
        except Exception as e:
            self.running = False
            self.error = True
            # Remove non-ascii characters in exception message (this can happen)
            printable = set(string.printable)
            self.status = ''.join(filter(lambda x: x in printable, str(e)))
            app.logger.error("Failed in task {}.".format(self.__class__.__name__), exc_info=e)
            return False
        else:
            self.completed = True
            self.running = False
            self.status = None
            return True

    def run(self):
        pass

    @property
    def status(self):
        return self._status

    @status.setter
    def status(self, val):
        self._status = val
        self.job.queue.put("status")


class Job(object):
    def __init__(self, worker, username, projectname, sample_filename,
            sample_file_object):
        self.worker = worker
        self.project_template_data = worker.project_template_data
        self.projectname = projectname
        # Get user object used as owner of objects created in LIMS. This also
        # checks that the username is valid in LIMS.
        self.user = lims.get_researchers(username=username)[0]
        self.sample_filename = sample_filename
        self.sample_file_object = sample_file_object
        self.samples = []
        self.lims_project = None
        self.lims_samples = []
        self.step_url = None    # Not used; may use in the future.
        self.queue = Mod_Queue.Queue()
        self.tasks = [
            getattr(sys.modules[__name__], task)(self)
            for task in self.project_template_data['tasks']
        ]

    def run(self):
        for task in self.tasks:
            if not task():
                break
        self.queue.put("shutdown")

    @property
    def running(self):
        return any(task.running for task in self.tasks)

    @property
    def error(self):
        return any(task.error for task in self.tasks)

    @property
    def completed(self):
        return all(task.completed for task in self.tasks)


def check_sample_list(samples):
    """Check sample list"""
    if not samples: raise ValueError("No samples found in sample file. Check the format.")
    names = [name for name, _, _ in samples]
    if len(set(names)) != len(samples):
        raise ValueError("Non-unique sample name(s): {}".format(
            list(set(name for name in names if names.count(name) > 1))
        ))
    if len(set(pos for _, pos, _ in samples)) != len(samples):
        raise ValueError("Duplicate well position(s) detected")
    if not all(c.isalnum() or c == "-"
            for name, _, _ in samples
            for c in name):
        raise ValueError("Invalid characters in sample name, A-Z, 0-9 and - allowed.")


class ReadMIKSampleFile(Task):
    """Read the sample file input (MIK)."""

    NAME = "Read sample file"

    def run(self):
        wb = load_workbook(self.job.sample_file_object, data_only=True)
        sheet = next(iter(wb))
        for coord, expect in zip(['A1','B1','C1'],
            [['Well'], ['Well Name'], ['E-gen', 'N-gen']]):
            if sheet[coord].value not in expect:
                raise ValueError("MIK file error: expected '{}' at {}, found '{}' instead.".format(
                            " or ".join(expect), coord, sheet[coord].value
                        ))
        additional_headers = []
        for acol in "DEFGHIJK":
            h = sheet["{}1".format(acol)].value
            if h:
                additional_headers.append(h.encode('ascii', errors='ignore'))
            else:
                break
        self.job.samples = []
        for row in sheet.iter_rows(min_row=2, max_col=3+len(additional_headers)):
            pos, name, ct = [c.value for c in row[0:3]]
            if not name: continue
            name = str(name)
            if name == "---": continue # Skip blank cells
            name = re.sub(r"[^A-Za-z0-9-]", "-", name)
            if additional_headers:
                ac = ";".join(
                    "{}={}".format(h, cell.value or '')
                    for h, cell in zip(additional_headers, row[3:])
                )
                udf_dict = {"Additional columns (MIK)": ac}
            else:
                udf_dict = {}
            try:
                if ct and ct != "No Cq":
                    udf_dict['Org. Ct value'] = float(ct)
            except ValueError:
                raise ValueError("Invalid Ct value '{}' at {}.".format(ct, row[2].coordinate))
            m = re.match(r"([A-H])(\d+)$", pos)
            if m and 1 <= int(m.group(2)) <= 12:
                self.job.samples.append((
                            name,
                            "{}:{}".format(m.group(1), m.group(2)),
                            udf_dict
                            ))
            else:
                raise ValueError("Invalid well position '{}'".format(pos))
        check_sample_list(self.job.samples)


class ReadFHISampleFile(Task):
    """Read the sample file input."""

    NAME = "Read sample file"

    def run(self):
        wb = load_workbook(self.job.sample_file_object, data_only=True)
        sheet = next(iter(wb))
        expected_headers = ['Position on plate', 'SampleID', 'Original ct-value']
        for header_row in range(1,6):
            if all(
                str(sheet[coord.format(header_row)].value).strip().lower() == expect.lower()
                for coord, expect in zip(['A{}','B{}','C{}'], expected_headers)
            ):
                break # Found it
        else: # This is if we didn't find the header column
            raise ValueError(
                "FHI file error: the expected headers {} were not found in the first {} lines.".format(
                            expected_headers, header_row
                            ))

        self.job.samples = []
        for row in sheet.iter_rows(min_row=header_row+1, max_col=3):
            pos, name_val, ct = [c.value for c in row]
            name = str(name_val)
            udf_dict = {}
            try:
                if ct:
                    udf_dict = {'Org. Ct value': float(ct)}
            except ValueError:
                raise ValueError("Invalid Ct value '{}' at {}.".format(ct, row[2].coordinate))
            if not pos or not name_val: # Blank lines after table are not the last ones
                break
            m = re.match(r"([A-H])(\d+)$", pos)
            if m and 1 <= int(m.group(2)) <= 12:
                self.job.samples.append((
                            name,
                            "{}:{}".format(m.group(1), m.group(2)),
                            udf_dict
                            ))
            else:
                raise ValueError("Invalid well position '{}'".format(pos))
        check_sample_list(self.job.samples)
        


class ReadFHIv2SampleFile(Task):
    """Read the sample file input."""

    NAME = "Read sample file (v2 csv file)"

    def run(self):
        try:
            # Note about BytesIO.getvalue(): In newer Python versions it is called getbuffer()
            # instead. It would be nice to use read() - but that returns an empty array.
            #text = self.job.sample_file_object.getbuffer().decode('utf-8').splitlines()
            # Also - different semantics of decoding characters - we can only support ASCII
            # but that's okay.
            text = bytes(self.job.sample_file_object.getvalue()).decode('utf-8-sig').splitlines()
        except UnicodeDecodeError as e:
            raise ValueError(
                "The file has incorrect format. Make sure to use a text-based format (CSV). "
                "Remove any special characters in the file. [Detailed error:" + str(e) + "]"
            )

        reader = csv.reader(text, delimiter=';')
        
        required_columns = set(['position', 'sequence id', 'ct', 'control name'])
        column_indices = {}

        for row in reader:
            is_blank_row = not any(row)
            if not column_indices: # Look for header
                if not is_blank_row:
                    lowercase_headers = [header.lower() for header in row]
                    column_indices = {header: i for i, header in enumerate(lowercase_headers)}
                    missing_columns = required_columns - set(lowercase_headers)
                    if missing_columns:
                        raise ValueError("The file is missing the following required column(s) "
                                        "(headers not found): " + ",".join(missing_columns))
                else:
                    continue # Keep skipping blank rows

            elif not is_blank_row: # Already have headers - read samples here
                udf_dict = {}
                name = row[column_indices['sequence id']]
                if not name: raise ValueError("Sample does not have a sequence ID")
                ctrl_name = row[column_indices['control name']]
                if ctrl_name:
                    udf_dict['Control Name'] = ctrl_name
                ct_str = row[column_indices['ct']]
                if ct_str:
                    try:
                        udf_dict['Org. Ct value'] = float(ct_str.replace(",","."))
                    except ValueError:
                        raise ValueError("Invalid Ct value '" + str(ct_str) + "'.")

                well_str = row[column_indices['position']]
                m = re.match(r"([A-H])(\d+)$", well_str)
                if m and 1 <= int(m.group(2)) <= 12:
                    self.job.samples.append((
                                name,
                                "{}:{}".format(m.group(1), m.group(2)),
                                udf_dict
                                ))
                else:
                    raise ValueError("Invalid well position '{}'".format(pos))
        
        if not column_indices:
            raise ValueError("Unable to find the column headers in the CSV file - invalid format. "
                            "Make sure to use a semicolon-delimited (;) CSV file.")


class CheckExistingProjects(Task):
    """Refuse to create project if one with the same name exists."""

    NAME = "Check for existing project"
    def run(self):
        projects = lims.get_projects(name=self.job.projectname)
        if projects:
            raise ValueError("A project named {0} already exists.".format(
                self.job.projectname))


class CreateProject(Task):
    NAME = "Create project"

    def run(self):
        self.job.lims_project = lims.create_project(
                name=self.job.projectname,
                researcher=self.job.user,
                open_date=datetime.date.today(),
                udf=self.job.project_template_data['project_fields']
                )
        

class UploadFile(Task):
    NAME = "Upload sample file"

    def run(self):
        # Would use the upstream file API, but it only supports a real on-disk file!
        gls = lims.glsstorage(self.job.lims_project, self.job.sample_filename)
        f_obj = gls.post()
        f_obj.upload(self.job.sample_file_object.getvalue())


class CreatePlateAndSamples(Task):
    NAME = "Create plate and samples"

    def run(self):
        plate = lims.create_container(type=lims.get_container_types('96 well plate')[0])
        for sample in self.job.samples:
            set_udfs = sample[2]
            if 'sample_fields' in self.job.project_template_data and self.job.project_template_data['sample_fields']:
                set_udfs.update(self.job.project_template_data['sample_fields'])
            lims_sample = lims.create_sample(
                    sample[0],
                    self.job.lims_project,
                    container=plate,
                    well=sample[1],
                    udf=set_udfs
            )
            self.job.lims_samples.append(lims_sample)
            

class AssignWorkflow(Task):
    NAME = "Assign to workflow"

    def run(self):
        artifacts = [sample.artifact for sample in self.job.lims_samples]
        workflows = lims.get_workflows(name=self.job.project_template_data['workflow'])
        if not workflows:
            raise ValueError("Specified workflow {0} does not exist.".format(
                self.job.project_template_data['workflow']))
        self.job.lims_workflow = workflows[0]
        lims.route_artifacts(artifacts, workflow_uri=self.job.lims_workflow.uri)


class ProjectWorker(object):
    """Worker, manages the trhead. Will run one Job, which is to import one project.
    
    (In version 1 project importer, the worker runs all jobs for a specific project type.)
    """

    def __init__(self, project_template_data):
        self.project_template_data = project_template_data
        self.job = None
        self.active = False
        self.rundate = None

    def start_job(self, username, password, project_title, sample_filename,
            sample_file_object):
        """Start an import task with the specified parameters.
        
        This function is not thread safe! Syncrhonization must be handled
        by the caller.
        """
        assert not self.active

        self.check_lims_credentials(username, password)
        self.job = Job(self, username, project_title, sample_filename,
                sample_file_object)
        thread = threading.Thread(target=self.run_job)
        self.active = True
        self.rundate = None
        thread.start()

    def check_lims_credentials(self, username, password):
        """Check supplied username and password. Throws an
        exception if they are incorrect, or if connection to
        LIMS fails."""

        uri = config.BASEURI.rstrip("/") +  '/api/'
        # Building header manually because of problem with unicode encoding
        bb = base64.b64encode(u"{}:{}".format(username,password).encode('utf-8'))
        encoded_auth = 'Basic ' + str(bb)
        r = requests.get(uri, headers={'Authorization': encoded_auth})
        if r.status_code not in [403, 200]:
            if r.status_code in [500]:
                raise RuntimeError("Internal server error")
            else:
                # Note: 403 is OK! It indicates that we have a valid password, but
                # are a Researcher user and thus not allowed to access the API. 
                # If the password is wrong, we will get a 401.
                raise LimsCredentialsError()

    def run_job(self):
        try:
            self.job.run()
        finally:
            self.active = False
            self.rundate = datetime.date.today()


# Start deveopment server if called on the command line
if __name__ == "__main__":
    app.debug = True
    app.run(host="0.0.0.0", port=5001, threaded=True)
